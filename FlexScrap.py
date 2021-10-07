def flex_scrap():
    from openpyxl import load_workbook, Workbook
    month = input("Scrap Month? ")
    shift_notes = load_workbook(filename = 'flex throughput nonsense\Verathon_FLEX_daily report -  '+month+' 2021.xlsx')
    flex_scrap_book = Workbook()
    sheet = flex_scrap_book.active

    header1 = ["Flex 2.0 Scrap","","","","","","","","","","Flex 2","","","flex 3"]
    sheet.append(header1)
    header2 = ["Date","Fallen Rims","Rim Loader","QA","Flex 1 Scrap", "Total", "Pcs Produced", "","", "Date", "Fallen Rims F2", "Fallen Rims F3", "Rim Loader F2", "Rim Loader F3", "QA F2","QA F3","Total","F2 Total","pc Produced","%",]
    sheet.append(header2)


    for sheet_ranges in shift_notes.worksheets:
        fallen_rims_f2 = 0
        fallen_rims_f3 = 0
        rim_loader_f2 = 0
        rim_loader_f3 = 0
        qa_f2 = 0
        qa_f3 = 0
        rims_made = 0
        flex_scrap_list = []
        flex_scrap_list.append(sheet_ranges['J3'].value) #date
        try:
            fallen_rims_f2 += sheet_ranges['Z83'].value / 0.012
        except:
            fallen_rims_f2 += 0
        try:
            fallen_rims_f2 += sheet_ranges['Z84'].value / 0.012
        except:
            fallen_rims_f2 += 0
        try:
            fallen_rims_f2 += sheet_ranges['Z85'].value / 0.012
        except:
            fallen_rims_f2 += 0
        flex_scrap_list.append(fallen_rims_f2)
        # fallen_rims_f3
        try:
            fallen_rims_f3 += sheet_ranges['Z90'].value / 0.012
        except:
            fallen_rims_f3 += 0
        try:
            fallen_rims_f3 += sheet_ranges['Z91'].value / 0.012
        except:
            fallen_rims_f3 += 0
        try:
            fallen_rims_f3 += sheet_ranges['Z92'].value / 0.012
        except:
            fallen_rims_f3 += 0
        flex_scrap_list.append(fallen_rims_f3)
        total_fallen_rims = fallen_rims_f3 + fallen_rims_f2
        flex_scrap_list.insert(1,total_fallen_rims)
        # rim_loader_f2
        try:
            rim_loader_f2 += sheet_ranges['AB83'].value / 0.012
        except:
            rim_loader_f2 += 0
        try:
            rim_loader_f2 += sheet_ranges['AB84'].value / 0.012
        except:
            rim_loader_f2 += 0
        try:
            rim_loader_f2 += sheet_ranges['AB85'].value / 0.012
        except:
            rim_loader_f2 += 0
        flex_scrap_list.append(rim_loader_f2)
        # rim_loader_f3
        try:
            rim_loader_f3 += sheet_ranges['AB90'].value / 0.012
        except:
            rim_loader_f3 += 0
        try:
            rim_loader_f3 += sheet_ranges['AB91'].value / 0.012
        except:
            rim_loader_f3 += 0
        try:
            rim_loader_f3 += sheet_ranges['AB92'].value / 0.012
        except:
            rim_loader_f3 += 0
        flex_scrap_list.append(rim_loader_f3)
        total_rim_loader = rim_loader_f3 + rim_loader_f2
        flex_scrap_list.insert(2, total_rim_loader)
        # qa_f2
        try:
            qa_f2 += sheet_ranges['AD83'].value / 0.012
        except:
            qa_f2 += 0
        try:
            qa_f2 += sheet_ranges['AD84'].value / 0.012
        except:
            qa_f2 += 0
        try:
            qa_f2 += sheet_ranges['AD85'].value / 0.012
        except:
            qa_f2 += 0
        flex_scrap_list.append(qa_f2)
        # qa_f3
        try:
            qa_f3 += sheet_ranges['AD90'].value / 0.012
        except:
            qa_f3 += 0
        try:
            qa_f3 += sheet_ranges['AD91'].value / 0.012
        except:
            qa_f3 += 0
        try:
            qa_f3 += sheet_ranges['AD92'].value / 0.012
        except:
            qa_f3 += 0

        flex_scrap_list.append(qa_f3)
        total_qa = qa_f3 + qa_f2
        flex_scrap_list.insert(3,total_qa)

        #total sum of all rims produced
        try:
            rims_made += sheet_ranges['Y53'].value * 32

        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y54'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y55'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y56'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y63'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y64'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y65'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y66'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y73'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y74'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y75'].value * 32
        except:
            rims_made += 0
        try:
            rims_made += sheet_ranges['Y76'].value * 32
        except:
            rims_made += 0
        flex_scrap_list.append(rims_made)
        f2_total_scrap = total_qa + total_fallen_rims + total_rim_loader - qa_f3
        flex_scrap_list.append(f2_total_scrap)
        #blank space for Flex 1 Scrap
        flex_scrap_list.insert(4,"")
        flex_scrap_list.insert(5, total_rim_loader + total_qa + total_fallen_rims)
        flex_scrap_list.insert(6, rims_made)
        flex_scrap_list.insert(7,"")
        flex_scrap_list.insert(8,"")
        flex_scrap_list.insert(9, sheet_ranges['J3'].value)


        sheet.append(flex_scrap_list)

    sheet.column_dimensions['A'].width = 22
    sheet.column_dimensions['B'].width = 9
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['D'].width = 9
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 22
    sheet.column_dimensions['K'].width = 9
    sheet.column_dimensions['L'].width = 9
    sheet.column_dimensions['M'].width = 9
    sheet.column_dimensions['N'].width = 9
    sheet.column_dimensions['O'].width = 9
    sheet.column_dimensions['P'].width = 9
    sheet.column_dimensions['Q'].width = 9
    sheet.column_dimensions['R'].width = 9
    sheet.column_dimensions['S'].width = 9
    sheet.delete_rows(3,2)
    flex_scrap_book.save('flex throughput nonsense\Flex Scrap '+month+'.xlsx')

    print("Flex Scrap "+month+".xlsx Created")