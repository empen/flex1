#Import Shit
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border

#Month Set to read monthly sheet
month = input("Throughput Month? ")

#Flex Throughput
shift_notes = load_workbook(filename = 'flex throughput nonsense\Verathon_FLEX_daily report -  '+month+' 2021.xlsx')
flex_cup_shift_notes = load_workbook(filename = 'flex throughput nonsense\Mixed Assemblies Daily report.xlsx')
current_tracking_book = Workbook()
sheet = current_tracking_book.active

#Header Creations
header1 = ["","","","","SCRAP","","","DOWNTIME","","","Packaging"]
sheet.append(header1)
header2 = ["Date","Shift","Line","Discs Production","Fallen Rim", "Rim Loader", "QA", "Axiom", "Flow Wrap", "Box Line", "SKU", "Qty", "SKU", "Qty"]
sheet.append(header2)
sheet.merge_cells('E1:G1')
sheet['E1'].fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
sheet.merge_cells('H1:J1')
sheet['H1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
sheet.merge_cells('K1:N1')
sheet['K1'].fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
sheet.freeze_panes = "A3"

# for sheet_ranges_cups in flex_cup_shift_notes.worksheets:



for sheet_ranges, sheet_ranges2 in shift_notes.worksheets, flex_cup_shift_notes.worksheets:
    midnights_f2_list = []
    days_f2_list = []
    afternoons_f2_list = []
    m_f2_cup_sum = 0
    d_f2_cup_sum = 0
    a_f2_cup_sum = 0

    midnights_f3_list = []
    days_f3_list = []
    afternoons_f3_list = []
    m_f3_cup_sum = 0
    d_f3_cup_sum = 0
    a_f3_cup_sum = 0

    midnights_cup_list = []
    days_cup_list = []
    afternoons_cup_list = []
    #not used yet
    m_cup_sum = 0
    d_cup_sum = 0
    a_cup_sum = 0
    #MIDNIGHT Flex 2 List
    midnights_f2_list.append(sheet_ranges['J3'].value) #date
    midnights_f2_list.append("Midnights") #midnight shift
    midnights_f2_list.append("F2") #flex 2
    try:
        midnights_f2_list.append(sheet_ranges['Z83'].value / 0.012) #fallen rimaroos scraop
        m_f2_cup_sum = m_f2_cup_sum + sheet_ranges['Z83'].value / 0.012
    except:
        midnights_f2_list.append(0)
    try:
        midnights_f2_list.append(sheet_ranges['AB83'].value / 0.012) #rim loader scrap
        m_f2_cup_sum = m_f2_cup_sum + sheet_ranges['AB83'].value / 0.012
    except:
        midnights_f2_list.append(0)
    try:
        midnights_f2_list.append(sheet_ranges['AD83'].value / 0.012) #QA Scrap
        m_f2_cup_sum = m_f2_cup_sum + sheet_ranges['AD83'].value / 0.012
    except:
        midnights_f2_list.append(0)
    try:
        midnights_f2_list.append(sheet_ranges['S83'].value)
    except:
        midnights_f2_list.append(0)
    try:
        midnights_f2_list.append(sheet_ranges['T83'].value)
    except:
        midnights_f2_list.append(0)
    try:
        midnights_f2_list.append(sheet_ranges['U83'].value)
    except:
        midnights_f2_list.append(0)
    #SKU #1 Run
    midnights_f2_list.append(sheet_ranges['T53'].value)
    try:
        midnights_f2_list.append(sheet_ranges['Y53'].value * 32)
        m_f2_cup_sum = m_f2_cup_sum + sheet_ranges['Y53'].value * 32
    except:
        midnights_f2_list.append(0)
    #SKU #2 Run
    midnights_f2_list.append(sheet_ranges['T54'].value)
    try:
        midnights_f2_list.append(sheet_ranges['Y54'].value * 32)
        m_f2_cup_sum = m_f2_cup_sum + sheet_ranges['Y54'].value * 32
    except:
        midnights_f2_list.append(0)
    midnights_f2_list.insert(3,m_f2_cup_sum)
    #DAYS Flex 2 List
    days_f2_list.append(sheet_ranges['J3'].value) #date
    days_f2_list.append("Days") #days shift
    days_f2_list.append("F2") #flex 2
    try:
        days_f2_list.append(sheet_ranges['Z84'].value / 0.012) #fallen rimaroos scrap
        d_f2_cup_sum = d_f2_cup_sum + sheet_ranges['Z84'].value / 0.012
    except:
        days_f2_list.append(0)
    try:
        days_f2_list.append(sheet_ranges['AB84'].value / 0.012) #rim loader scrap
        d_f2_cup_sum = d_f2_cup_sum + sheet_ranges['AB84'].value / 0.012
    except:
        days_f2_list.append(0)
    try:
        days_f2_list.append(sheet_ranges['AD84'].value / 0.012) #QA Scrap
        d_f2_cup_sum = d_f2_cup_sum + sheet_ranges['AD84'].value / 0.012
    except:
        days_f2_list.append(0)
    try:
        days_f2_list.append(sheet_ranges['S84'].value)
    except:
        days_f2_list.append(0)
    try:
        days_f2_list.append(sheet_ranges['T84'].value)
    except:
        days_f2_list.append(0)
    try:
        days_f2_list.append(sheet_ranges['U84'].value)
    except:
        days_f2_list.append(0)
    #SKU #1 Run
    days_f2_list.append(sheet_ranges['T63'].value)
    try:
        days_f2_list.append(sheet_ranges['Y63'].value * 32)
        d_f2_cup_sum = d_f2_cup_sum + sheet_ranges['Y63'].value * 32
    except:
        days_f2_list.append(0)
    #SKU #2 Run
    days_f2_list.append(sheet_ranges['T64'].value)
    try:
        days_f2_list.append(sheet_ranges['Y64'].value * 32)
        d_f2_cup_sum = d_f2_cup_sum + sheet_ranges['Y64'].value * 32
    except:
        days_f2_list.append(0)
    days_f2_list.insert(3,d_f2_cup_sum)

    #Afternoons Flex 2 List
    afternoons_f2_list.append(sheet_ranges['J3'].value) #date
    afternoons_f2_list.append("Afternoons") #afternoons shift
    afternoons_f2_list.append("F2") #flex 2
    try:
        afternoons_f2_list.append(sheet_ranges['Z85'].value / 0.012) #fallen rimaroos scraop
        a_f2_cup_sum = a_f2_cup_sum + sheet_ranges['Z85'].value / 0.012
    except:
        afternoons_f2_list.append(0)
    try:
        afternoons_f2_list.append(sheet_ranges['AB85'].value / 0.012) #rim loader scrap
        a_f2_cup_sum = a_f2_cup_sum + sheet_ranges['AB85'].value / 0.012
    except:
        afternoons_f2_list.append(0)
    try:
        afternoons_f2_list.append(sheet_ranges['AD85'].value / 0.012) #QA Scrap
        a_f2_cup_sum = a_f2_cup_sum + sheet_ranges['AD85'].value / 0.012
    except:
        afternoons_f2_list.append(0)
    try:
        afternoons_f2_list.append(sheet_ranges['S85'].value)
    except:
        afternoons_f2_list.append(0)
    try:
        afternoons_f2_list.append(sheet_ranges['T85'].value)
    except:
        afternoons_f2_list.append(0)
    try:
        afternoons_f2_list.append(sheet_ranges['U85'].value)
    except:
        afternoons_f2_list.append(0)
    #SKU #1 Run
    afternoons_f2_list.append(sheet_ranges['T73'].value)
    try:
        afternoons_f2_list.append(sheet_ranges['Y73'].value * 32)
        a_f2_cup_sum = a_f2_cup_sum + sheet_ranges['Y73'].value * 32
    except:
        afternoons_f2_list.append(0)
    #SKU #2 Run
    afternoons_f2_list.append(sheet_ranges['T74'].value)
    try:
        afternoons_f2_list.append(sheet_ranges['Y74'].value * 32)
        a_f2_cup_sum = a_f2_cup_sum + sheet_ranges['Y74'].value * 32
    except:
        afternoons_f2_list.append(0)
    afternoons_f2_list.insert(3,a_f2_cup_sum)

    #MIDNIGHT Flex 3 List
    midnights_f3_list.append(sheet_ranges['J3'].value)  # date
    midnights_f3_list.append("Midnights")  # midnight shift
    midnights_f3_list.append("F3")  # flex 3
    try:
        midnights_f3_list.append(sheet_ranges['Z90'].value / 0.012)  # fallen rimaroos scraop
        m_f3_cup_sum = m_f3_cup_sum + sheet_ranges['Z90'].value / 0.012
    except:
        midnights_f3_list.append(0)
    try:
        midnights_f3_list.append(sheet_ranges['AB90'].value / 0.012)  # rim loader scrap
        m_f3_cup_sum = m_f3_cup_sum + sheet_ranges['AB90'].value / 0.012
    except:
        midnights_f3_list.append(0)
    try:
        midnights_f3_list.append(sheet_ranges['AD90'].value / 0.012)  # QA Scrap
        m_f3_cup_sum = m_f3_cup_sum + sheet_ranges['AD90'].value / 0.012
    except:
        midnights_f3_list.append(0)
    try:
        midnights_f3_list.append(sheet_ranges['S90'].value)
    except:
        midnights_f3_list.append(0)
    try:
        midnights_f3_list.append(sheet_ranges['T90'].value)
    except:
        midnights_f3_list.append(0)
    try:
        midnights_f3_list.append(sheet_ranges['U90'].value)
    except:
        midnights_f3_list.append(0)
    # SKU #1 Run
    midnights_f3_list.append(sheet_ranges['T55'].value)
    try:
        midnights_f3_list.append(sheet_ranges['Y55'].value * 32)
        m_f3_cup_sum = m_f3_cup_sum + sheet_ranges['Y55'].value * 32
    except:
        midnights_f3_list.append(0)
    # SKU #2 Run
    midnights_f3_list.append(sheet_ranges['T56'].value)
    try:
        midnights_f3_list.append(sheet_ranges['Y56'].value * 32)
        m_f3_cup_sum = m_f3_cup_sum + sheet_ranges['Y56'].value * 32
    except:
        midnights_f3_list.append(0)
    midnights_f3_list.insert(3, m_f3_cup_sum)

    #DAYS Flex 3 List
    days_f3_list.append(sheet_ranges['J3'].value) #date
    days_f3_list.append("Days") #days shift
    days_f3_list.append("F3") #flex 3
    try:
        days_f3_list.append(sheet_ranges['Z91'].value / 0.012) #fallen rimaroos scraop
        d_f3_cup_sum = d_f3_cup_sum + sheet_ranges['Z91'].value / 0.012
    except:
        days_f3_list.append(0)
    try:
        days_f3_list.append(sheet_ranges['AB91'].value / 0.012) #rim loader scrap
        d_f3_cup_sum = d_f3_cup_sum + sheet_ranges['AB91'].value / 0.012
    except:
        days_f3_list.append(0)
    try:
        days_f3_list.append(sheet_ranges['AD91'].value / 0.012) #QA Scrap
        d_f3_cup_sum = d_f3_cup_sum + sheet_ranges['AD91'].value / 0.012
    except:
        days_f3_list.append(0)
    try:
        days_f3_list.append(sheet_ranges['S91'].value)
    except:
        days_f3_list.append(0)
    try:
        days_f3_list.append(sheet_ranges['T91'].value)
    except:
        days_f3_list.append(0)
    try:
        days_f3_list.append(sheet_ranges['U91'].value)
    except:
        days_f3_list.append(0)
    #SKU #1 Run
    days_f3_list.append(sheet_ranges['T65'].value)
    try:
        days_f3_list.append(sheet_ranges['Y65'].value * 32)
        d_f3_cup_sum = d_f3_cup_sum + sheet_ranges['Y65'].value * 32
    except:
        days_f3_list.append(0)
    #SKU #2 Run
    days_f3_list.append(sheet_ranges['T66'].value)
    try:
        days_f3_list.append(sheet_ranges['Y66'].value * 32)
        d_f3_cup_sum = d_f3_cup_sum + sheet_ranges['Y66'].value * 32
    except:
        days_f3_list.append(0)
    days_f3_list.insert(3,d_f3_cup_sum)

    #AFTERNOONS Flex 3 List
    afternoons_f3_list.append(sheet_ranges['J3'].value) #date
    afternoons_f3_list.append("Afternoons") #afternoons shift
    afternoons_f3_list.append("F3") #flex 3
    try:
        afternoons_f3_list.append(sheet_ranges['Z92'].value / 0.012) #fallen rimaroos scraop
        a_f3_cup_sum = a_f3_cup_sum + sheet_ranges['Z92'].value / 0.012
    except:
        afternoons_f3_list.append(0)
    try:
        afternoons_f3_list.append(sheet_ranges['AB92'].value / 0.012) #rim loader scrap
        a_f3_cup_sum = a_f3_cup_sum + sheet_ranges['AB92'].value / 0.012
    except:
        afternoons_f3_list.append(0)
    try:
        afternoons_f3_list.append(sheet_ranges['AD92'].value / 0.012) #QA Scrap
        a_f3_cup_sum = a_f3_cup_sum + sheet_ranges['AD92'].value / 0.012
    except:
        afternoons_f3_list.append(0)
    try:
        afternoons_f3_list.append(sheet_ranges['S92'].value)
    except:
        afternoons_f3_list.append(0)
    try:
        afternoons_f3_list.append(sheet_ranges['T92'].value)
    except:
        afternoons_f3_list.append(0)
    try:
        afternoons_f3_list.append(sheet_ranges['U92'].value)
    except:
        afternoons_f3_list.append(0)
    #SKU #1 Run
    afternoons_f3_list.append(sheet_ranges['T75'].value)
    try:
        afternoons_f3_list.append(sheet_ranges['Y75'].value * 32)
        a_f3_cup_sum = a_f3_cup_sum + sheet_ranges['Y75'].value * 32
    except:
        afternoons_f3_list.append(0)
    #SKU #2 Run
    afternoons_f3_list.append(sheet_ranges['T76'].value)
    try:
        afternoons_f3_list.append(sheet_ranges['Y76'].value * 32)
        a_f3_cup_sum = a_f3_cup_sum + sheet_ranges['Y76'].value * 32
    except:
        afternoons_f3_list.append(0)
    afternoons_f3_list.insert(3,a_f3_cup_sum)

    #FLEX CUP LIST START--------------- CURRENTLY BLANK-----------
    #MIDNIGHT Cup List
    midnights_cup_list.append(sheet_ranges['J3'].value)  # date
    midnights_cup_list.append("Midnights")  # midnight shift
    midnights_cup_list.append("Cup")  # flex Cup

    #DAYS Cup List
    days_cup_list.append(sheet_ranges['J3'].value)  # date
    days_cup_list.append("Days")  # midnight shift
    days_cup_list.append("Cup")  # flex Cup
    days_cup_list.append(sheet_ranges2['R45'].vale * 18)

    #AFTERNOONS Cup List
    afternoons_cup_list.append(sheet_ranges['J3'].value)  # date
    afternoons_cup_list.append("Afternoon")  # midnight shift
    afternoons_cup_list.append("Cup")  # flex Cup



    #FLEX CUP LIST END--------------- CURRENTLY BLANK-----------

    #BARF LISTS ONTO SHEET
    sheet.append(midnights_f2_list)
    sheet.append(days_f2_list)
    sheet.append(afternoons_f2_list)

    sheet.append(midnights_f3_list)
    sheet.append(days_f3_list)
    sheet.append(afternoons_f3_list)

    sheet.append(midnights_cup_list)
    sheet.append(days_cup_list)
    sheet.append(afternoons_cup_list)

    #Border Formatting


#unugly columns
sheet.column_dimensions['A'].width = 22
sheet.column_dimensions['B'].width = 13
sheet.column_dimensions['C'].width = 6
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 10
sheet.column_dimensions['H'].width = 10
sheet.column_dimensions['I'].width = 10
sheet.column_dimensions['J'].width = 10
sheet.column_dimensions['K'].width = 16
sheet.column_dimensions['L'].width = 10
sheet.column_dimensions['M'].width = 16
sheet.column_dimensions['N'].width = 10

#format Date/Time


#delete garbo from first 2 sheets, save and tell you it's done.
sheet.delete_rows(3,18)
current_tracking_book.save('flex throughput nonsense\Flex Throughput ' +month+' test.xlsx')

print("Flex Throughput "+month+" test.xlsx Created")